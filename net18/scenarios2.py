import numpy as np
import openpyxl
import pandas as pd
import os
from case118.dataset import Dataset
from torch.utils.data import DataLoader
import torch
from case118.networks import ANN


def get_data_by_scenario_and_case(scenario, case):
    assert case > 0
    assert scenario > 0

    file_excel = f'./validation_data/Scenario{scenario}.xlsx'
    xls = pd.ExcelFile(file_excel)
    scheda = xls.sheet_names[case-1]
    dati_scheda = pd.read_excel(file_excel, sheet_name=scheda)
    valori_veri_misure = dati_scheda['Valori veri misure'].tolist()
    misure = dati_scheda['Misure'].tolist()
    valori_veri_stima = dati_scheda['Valori veri stima'].tolist()
    valori_stimati = dati_scheda['Valori stimati'].tolist()
    x = np.hstack((
        [list(-1. * np.asarray(valori_veri_misure[5:22]))], [list(-1. * np.asarray(valori_veri_misure[22:39]))], [valori_veri_misure[:5]], [valori_veri_misure[39:46]], [valori_veri_misure[46:]]
    ))

    x_hat = np.hstack((
        [list(-1. * np.asarray(misure[5:22]))], [list(-1. * np.asarray(misure[22:39]))], [misure[:5]], [misure[39:46]], [misure[46:]]
    ))
    y = [valori_veri_stima[:18]]
    y_hat = [valori_stimati[:18]]

    sens_tab = pd.read_excel(file_excel, sheet_name='Sensitivity')
    sens = []


    # sens = np.hstack((s1_sens_pi, s1_sens_qi, s1_sens_vmpu, s1_sens_pf, s1_sens_qf))
    return x, x_hat, y, y_hat, sens

def report_results_by_scenario_and_case(model, device='cpu'):
    scenarios = 5
    cases = 3

    model = ANN(53, 1500, 18)
    model.load_state_dict(torch.load("model_net18_53.pth"))
    model.eval()
    model = model.to(device)


    for scenario in range(scenarios):
        scenario += 1

        workbook = openpyxl.Workbook()
        for case in range(cases):
            case += 1
            workbook.create_sheet(title=f'Sheet{case}')
            sheet = workbook[f'Sheet{case}']
            data = list()
            y, pred = get_results(scenario, case, model)
            data.append(np.asarray(range(18)))
            data.append(y.ravel())
            data.append(pred.ravel())
            #data = pd.DataFrame(np.asarray(data))
            for i, entry in enumerate(data):
                for j, value in enumerate(entry):
                    sheet.cell(row=j+1, column=i+1, value=value)
        workbook.save(f'results_scenario{scenario}.xlsx')


def get_results(scenario, case, model, device='cpu'):
    s_c_data = get_data_by_scenario_and_case(scenario, case)
    x = s_c_data[0]
    x_hat = s_c_data[1]
    y = s_c_data[2]
    y_hat = s_c_data[3]
    sens = s_c_data[4]

    test_data = Dataset(x_hat, y)
    test_dataloader = DataLoader(test_data, batch_size=len(test_data))

    for batch, (X, y) in enumerate(test_dataloader):
        X, y = X.to(device), y.to(device)
        pred = model(X)
        y = y.cpu().detach().numpy()
        pred = pred.cpu().detach().numpy()
        return y, pred



def report_results_on_validation_files(model, device='cpu'):
    scenarios = 5
    cases = 3

    model = ANN(53, 1580, 18, dropout=0.2302727528297988)
    model.load_state_dict(torch.load("model_net18_53.pth"))
    model.eval()
    model = model.to(device)


    for scenario in range(scenarios):
        scenario += 1

        workbook = openpyxl.load_workbook(f'./validation_data/Scenario{scenario}.xlsx')
        for case in range(cases):
            case += 1
            sheet = workbook[f'Caso{case}']
            data = list()
            _, pred = get_results(scenario, case, model)

            data.append(pred.ravel())
            sheet.cell(row=1, column=7, value='Valori stimati MLP')
            for i, entry in enumerate(data):
                for j, value in enumerate(entry):
                    sheet.cell(row=j+2, column=7, value=value)
        workbook.save(f'./validation_data/Scenario{scenario}.xlsx')

def report_preds_on_validation_files(preds, column_index, model_name, scenario, case=1):

    workbook = openpyxl.load_workbook(f'./validation_data/Scenario{scenario}.xlsx')

    sheet = workbook[f'Caso{case}']
    data = list()


    data.append(preds.ravel())
    sheet.cell(row=1, column=column_index, value=f'Valori stimati {model_name}')
    for i, entry in enumerate(data):
        for j, value in enumerate(entry):
            sheet.cell(row=j+2, column=column_index, value=value)
    workbook.save(f'./validation_data/Scenario{scenario}.xlsx')

report_results_on_validation_files(model=None)

#report_results_by_scenario_and_case(model=None)

#get_data_by_scenario_and_case(1,1)


